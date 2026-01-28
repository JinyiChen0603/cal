# -*- coding: utf-8 -*-
"""
对比Excel中的老师-ID和CSV中的老师-ID，找出差异
"""

import openpyxl
import csv

def parse_teachers(s):
    """解析老师列表"""
    if not s or not str(s).strip():
        return []
    return [x.strip() for x in str(s).split() if x.strip()]

def main():
    xlsx_path = "judge.xlsx"
    csv_path = "salary_detail.csv"
    
    print("="*80)
    print("对比Excel和CSV中的老师-ID")
    print("="*80)
    
    # 1. 从Excel提取所有唯一的老师-ID
    print("\n[1/3] 从Excel提取所有老师-ID...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    excel_teacher_ids = set()
    excel_teacher_details = {}  # 记录每个老师出现在哪些题目中
    
    for row in range(2, ws.max_row + 1):
        problem_id = ws.cell(row, 1).value
        
        # C列 passed_users
        passed = parse_teachers(ws.cell(row, 3).value)
        for t in passed:
            excel_teacher_ids.add(t)
            if t not in excel_teacher_details:
                excel_teacher_details[t] = []
            excel_teacher_details[t].append((problem_id, 'passed'))
        
        # D列 failed_users
        failed = parse_teachers(ws.cell(row, 4).value)
        for t in failed:
            excel_teacher_ids.add(t)
            if t not in excel_teacher_details:
                excel_teacher_details[t] = []
            excel_teacher_details[t].append((problem_id, 'failed'))
    
    wb.close()
    print(f"   Excel中找到 {len(excel_teacher_ids)} 个唯一的老师-ID")
    
    # 2. 从CSV读取所有老师-ID
    print("\n[2/3] 从CSV读取所有老师-ID...")
    csv_teacher_ids = set()
    
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            csv_teacher_ids.add(row['老师'])
    
    print(f"   CSV中找到 {len(csv_teacher_ids)} 个老师-ID")
    
    # 3. 找出差异
    print("\n[3/3] 对比差异...")
    missing = sorted(excel_teacher_ids - csv_teacher_ids)  # Excel有但CSV没有
    extra = sorted(csv_teacher_ids - excel_teacher_ids)    # CSV有但Excel没有
    
    print("\n" + "="*80)
    print("对比结果")
    print("="*80)
    
    print(f"\nExcel中有: {len(excel_teacher_ids)} 个老师-ID")
    print(f"CSV中有:   {len(csv_teacher_ids)} 个老师-ID")
    print(f"差异:      {len(excel_teacher_ids) - len(csv_teacher_ids)} 条")
    
    if missing:
        print(f"\n" + "="*80)
        print(f"Excel中有但CSV中缺失的老师-ID: {len(missing)} 个")
        print("="*80)
        
        for teacher_id in missing:
            appearances = excel_teacher_details[teacher_id]
            print(f"\n老师-ID: {teacher_id}")
            print(f"  出现次数: {len(appearances)}")
            print(f"  题目详情:")
            for problem_id, status in appearances:
                print(f"    - 题目 {problem_id} ({status})")
    else:
        print("\n✓ Excel中的所有老师-ID都在CSV中")
    
    if extra:
        print(f"\n" + "="*80)
        print(f"CSV中有但Excel中不存在的老师-ID: {len(extra)} 个")
        print("="*80)
        
        for teacher_id in extra:
            print(f"  - {teacher_id}")
    else:
        print("\n✓ CSV中没有多余的老师-ID")
    
    # 检查重复使用的老师-ID
    print("\n" + "="*80)
    print("检查重复使用的老师-ID")
    print("="*80)
    
    repeated_teachers = []
    for teacher_id, appearances in excel_teacher_details.items():
        if len(appearances) > 1:
            repeated_teachers.append((teacher_id, appearances))
    
    if repeated_teachers:
        print(f"\n发现 {len(repeated_teachers)} 个老师-ID被重复使用（评审了多个题目）:")
        total_repeats = 0
        for teacher_id, appearances in sorted(repeated_teachers):
            print(f"\n老师-ID: {teacher_id}")
            print(f"  评审次数: {len(appearances)}")
            print(f"  题目详情:")
            for problem_id, status in appearances:
                print(f"    - 题目 {problem_id} ({status})")
            total_repeats += len(appearances) - 1  # 减1是因为第一次不算重复
        
        print(f"\n总重复次数: {total_repeats}")
        print(f"计算验证: {len(excel_teacher_ids)} 个唯一ID + {total_repeats} 次重复 = {len(excel_teacher_ids) + total_repeats} 条总记录")
    else:
        print("\n✓ 没有重复使用的老师-ID，每个ID只评审了1个题目")
    
    # 汇总
    print("\n" + "="*80)
    print("问题汇总")
    print("="*80)
    if missing:
        print(f"\n可能原因：这 {len(missing)} 个老师-ID在生成CSV时被跳过了")
        print("建议检查 calc_salary_from_judge.py 中的数据处理逻辑")
        print("特别是这些老师所在行的K列（正确金额）和N列（错误金额）是否有异常")
    else:
        print("\n没有发现缺失的老师-ID，数据一致！")

if __name__ == "__main__":
    main()
