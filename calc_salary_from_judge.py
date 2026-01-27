# -*- coding: utf-8 -*-
"""
从 judge.xlsx 按每个格子的颜色判断「正确/错误」，汇总每位老师的正确、错误所得金
"""

from pathlib import Path
from collections import defaultdict
import openpyxl

def has_color(cell):
    """
    判断单个格子是否有浅蓝/浅绿填充
    根据实际文件，有颜色的格子特征：
    - patternType == 'solid'
    - theme == 8 (浅蓝) 或 theme == 9 (浅绿)
    """
    if not cell or not cell.fill:
        return False
    
    # 检查是否为实心填充
    if cell.fill.patternType != 'solid':
        return False
    
    # 检查主题颜色
    if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'theme'):
        theme = cell.fill.fgColor.theme
        # theme 8 = 浅蓝, theme 9 = 浅绿
        if theme in [8, 9]:
            return True
    
    return False

def parse_teachers(s):
    """把 '彭海航-1 李长葳-51 ...' 拆成 ['彭海航-1','李长葳-51',...]"""
    if not s or not str(s).strip():
        return []
    return [x.strip() for x in str(s).split() if x.strip()]

def extract_teacher_name(teacher_id):
    """从 '孙林-251' 提取出 '孙林'"""
    if '-' in teacher_id:
        return teacher_id.rsplit('-', 1)[0]  # 按最后一个 - 分割，取前面部分
    return teacher_id

def main():
    base = Path(__file__).resolve().parent
    xlsx_path = base / "judge.xlsx"
    
    print(f"正在读取文件: {xlsx_path}")
    # data_only=True 读取公式的计算结果而不是公式本身
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 列号：C=passed_users(3), D=failed_users(4), K=决定正确金额/每人(11), N=决定错误金额/每人(14)
    col_passed = 3
    col_failed = 4
    col_correct_per = 11   # K列
    col_wrong_per   = 14   # N列

    total_detail = defaultdict(lambda: [0.0, 0.0])   # 老师-ID -> [正确总金, 错误总金]
    total_summary = defaultdict(lambda: [0.0, 0.0])  # 老师姓名 -> [正确总金, 错误总金]

    print(f"开始处理 {ws.max_row - 1} 行数据...")
    
    for row in range(2, ws.max_row + 1):
        correct_per = ws.cell(row=row, column=col_correct_per).value
        wrong_per   = ws.cell(row=row, column=col_wrong_per).value
        
        try:
            correct_per = float(correct_per) if correct_per is not None else 0.0
            wrong_per   = float(wrong_per)   if wrong_per   is not None else 0.0
        except (TypeError, ValueError):
            continue

        # 处理 passed_users 格子（C列）
        cell_passed = ws.cell(row=row, column=col_passed)
        passed_list = parse_teachers(cell_passed.value)
        if has_color(cell_passed):
            # 有颜色 → 判断正确
            for t in passed_list:
                teacher_name = extract_teacher_name(t)
                total_detail[t][0] += correct_per       # 详细记录
                total_summary[teacher_name][0] += correct_per  # 汇总记录
        else:
            # 没颜色 → 判断错误
            for t in passed_list:
                teacher_name = extract_teacher_name(t)
                total_detail[t][1] += wrong_per
                total_summary[teacher_name][1] += wrong_per

        # 处理 failed_users 格子（D列）
        cell_failed = ws.cell(row=row, column=col_failed)
        failed_list = parse_teachers(cell_failed.value)
        if has_color(cell_failed):
            # 有颜色 → 判断正确
            for t in failed_list:
                teacher_name = extract_teacher_name(t)
                total_detail[t][0] += correct_per
                total_summary[teacher_name][0] += correct_per
        else:
            # 没颜色 → 判断错误
            for t in failed_list:
                teacher_name = extract_teacher_name(t)
                total_detail[t][1] += wrong_per
                total_summary[teacher_name][1] += wrong_per

    wb.close()

    # 输出到终端
    print("\n" + "="*80)
    print(f"处理完成！详细记录 {len(total_detail)} 条，汇总 {len(total_summary)} 位老师")
    print("="*80)
    
    # 写入详细文件
    detail_path = base / "salary_detail.csv"
    with open(detail_path, 'w', encoding='utf-8-sig', newline='') as f:
        f.write("老师,回答正确所得金,回答错误所得金,所得金合计\n")
        for t in sorted(total_detail.keys()):
            c, w = total_detail[t]
            f.write(f"{t},{c:.2f},{w:.2f},{c+w:.2f}\n")
    
    print(f"\n详细记录已写入: {detail_path}")
    
    # 写入汇总文件
    summary_path = base / "salary_summary.csv"
    with open(summary_path, 'w', encoding='utf-8-sig', newline='') as f:
        f.write("老师,回答正确所得金,回答错误所得金,所得金合计\n")
        for t in sorted(total_summary.keys()):
            c, w = total_summary[t]
            f.write(f"{t},{c:.2f},{w:.2f},{c+w:.2f}\n")
    
    print(f"汇总数据已写入: {summary_path}")
    print(f"\n详细记录: {len(total_detail)} 条")
    print(f"汇总老师: {len(total_summary)} 位")

if __name__ == "__main__":
    main()
