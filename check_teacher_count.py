# -*- coding: utf-8 -*-
"""
检查每个题目的评审老师数量是否为5位
"""

import openpyxl

def parse_teachers(s):
    """解析老师列表"""
    if not s or not str(s).strip():
        return []
    return [x.strip() for x in str(s).split() if x.strip()]

def main():
    xlsx_path = "judge.xlsx"
    
    print("="*80)
    print("检查每个题目的评审老师数量")
    print("="*80)
    
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    problems_with_issues = []
    total_problems = 0
    
    for row in range(2, ws.max_row + 1):
        problem_id = ws.cell(row, 1).value  # A列：题号/import_id
        
        # C列 passed_users
        cell_c = ws.cell(row, 3)
        passed_teachers = parse_teachers(cell_c.value)
        
        # D列 failed_users
        cell_d = ws.cell(row, 4)
        failed_teachers = parse_teachers(cell_d.value)
        
        # 合并所有老师
        all_teachers = passed_teachers + failed_teachers
        teacher_count = len(all_teachers)
        
        total_problems += 1
        
        # 检查是否为5位老师
        if teacher_count != 5:
            problems_with_issues.append({
                'row': row,
                'problem_id': problem_id,
                'count': teacher_count,
                'passed': passed_teachers,
                'failed': failed_teachers
            })
    
    wb.close()
    
    # 输出结果
    print(f"\n总共检查了 {total_problems} 个题目")
    print(f"其中 {len(problems_with_issues)} 个题目的老师数量不等于5\n")
    
    if problems_with_issues:
        print("不满5位老师的题目详情：")
        print("-" * 80)
        
        for item in problems_with_issues:
            print(f"\n题目ID: {item['problem_id']} (第{item['row']}行)")
            print(f"  实际老师数: {item['count']}")
            print(f"  Passed ({len(item['passed'])}人): {', '.join(item['passed']) if item['passed'] else '无'}")
            print(f"  Failed ({len(item['failed'])}人): {', '.join(item['failed']) if item['failed'] else '无'}")
        
        print("\n" + "="*80)
        print("汇总：异常题目ID列表")
        print("="*80)
        for item in problems_with_issues:
            print(f"{item['problem_id']}: {item['count']}位老师")
    else:
        print("\n✓ 所有题目都有5位老师！")
    
    # 额外统计：计算应该有多少条记录
    expected_records = total_problems * 5
    print(f"\n理论上应该有的记录数: {total_problems} 题目 × 5 老师 = {expected_records} 条")

if __name__ == "__main__":
    main()
