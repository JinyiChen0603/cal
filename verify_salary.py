# -*- coding: utf-8 -*-
"""
精确验证：逐条对比 salary_detail.csv 与 judge.xlsx 原始数据
"""

from pathlib import Path
from collections import defaultdict
import openpyxl
import csv

def has_color(cell):
    """判断单个格子是否有浅蓝/浅绿填充"""
    if not cell or not cell.fill:
        return False
    if cell.fill.patternType != 'solid':
        return False
    if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'theme'):
        theme = cell.fill.fgColor.theme
        if theme in [8, 9]:
            return True
    return False

def parse_teachers(s):
    """解析老师列表"""
    if not s or not str(s).strip():
        return []
    return [x.strip() for x in str(s).split() if x.strip()]

def extract_teacher_name(teacher_id):
    """从 '孙林-251' 提取出 '孙林'"""
    if '-' in teacher_id:
        return teacher_id.rsplit('-', 1)[0]
    return teacher_id

def main():
    base = Path(__file__).resolve().parent
    xlsx_path = base / "judge.xlsx"
    detail_path = base / "salary_detail.csv"
    
    print("="*80)
    print("精确验证：逐条对比 detail 与原始数据")
    print("="*80)
    
    # 1. 读取 detail.csv 的每条记录
    print("\n[1/3] 读取 salary_detail.csv...")
    detail_records = []
    with open(detail_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            detail_records.append({
                'teacher_id': row['老师'],
                'correct': float(row['回答正确所得金']),
                'wrong': float(row['回答错误所得金']),
                'total': float(row['所得金合计'])
            })
    print(f"   读取 {len(detail_records)} 条记录")
    
    # 2. 读取 Excel，建立"老师-ID"到题目的映射，同时统计汇总
    print("\n[2/3] 分析 judge.xlsx，建立映射关系...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    # teacher_id -> [(题号, 列名, 是否有颜色, correct_per, wrong_per), ...]
    teacher_mapping = defaultdict(list)
    # 老师姓名 -> [正确金额, 错误金额, 正确题数, 错误题数]
    excel_summary = defaultdict(lambda: [0.0, 0.0, 0, 0])
    
    for row in range(2, ws.max_row + 1):
        problem_id = ws.cell(row, 1).value  # A列：题号
        correct_per = ws.cell(row, 11).value  # K列
        wrong_per = ws.cell(row, 14).value    # N列
        
        try:
            correct_per = float(correct_per) if correct_per else 0.0
            wrong_per = float(wrong_per) if wrong_per else 0.0
        except:
            continue
        
        # C列 (passed_users)
        cell_c = ws.cell(row, 3)
        c_has_color = has_color(cell_c)
        for t in parse_teachers(cell_c.value):
            teacher_mapping[t].append({
                'problem_id': problem_id,
                'column': 'C(passed)',
                'has_color': c_has_color,
                'correct_per': correct_per,
                'wrong_per': wrong_per,
                'row': row
            })
            # 同时累加汇总统计
            teacher_name = extract_teacher_name(t)
            if c_has_color:
                excel_summary[teacher_name][0] += correct_per
                excel_summary[teacher_name][2] += 1
            else:
                excel_summary[teacher_name][1] += wrong_per
                excel_summary[teacher_name][3] += 1
        
        # D列 (failed_users)
        cell_d = ws.cell(row, 4)
        d_has_color = has_color(cell_d)
        for t in parse_teachers(cell_d.value):
            teacher_mapping[t].append({
                'problem_id': problem_id,
                'column': 'D(failed)',
                'has_color': d_has_color,
                'correct_per': correct_per,
                'wrong_per': wrong_per,
                'row': row
            })
            # 同时累加汇总统计
            teacher_name = extract_teacher_name(t)
            if d_has_color:
                excel_summary[teacher_name][0] += correct_per
                excel_summary[teacher_name][2] += 1
            else:
                excel_summary[teacher_name][1] += wrong_per
                excel_summary[teacher_name][3] += 1
    
    wb.close()
    print(f"   分析完成，找到 {len(teacher_mapping)} 个老师-ID")
    
    # 3. 逐条对比验证
    print("\n[3/3] 逐条验证...")
    errors = []
    
    for record in detail_records:
        teacher_id = record['teacher_id']
        csv_correct = record['correct']
        csv_wrong = record['wrong']
        
        # 从 Excel 重新计算这个老师-ID
        if teacher_id not in teacher_mapping:
            errors.append({
                'teacher_id': teacher_id,
                'error': 'CSV中存在但Excel中找不到此老师',
                'csv_data': record
            })
            continue
        
        excel_correct = 0.0
        excel_wrong = 0.0
        details = []
        
        for item in teacher_mapping[teacher_id]:
            if item['has_color']:
                excel_correct += item['correct_per']
                details.append(f"题{item['problem_id']}行{item['row']} {item['column']} 有颜色 +{item['correct_per']:.2f}(正确)")
            else:
                excel_wrong += item['wrong_per']
                details.append(f"题{item['problem_id']}行{item['row']} {item['column']} 无颜色 +{item['wrong_per']:.2f}(错误)")
        
        # 对比（允许0.01的浮点误差）
        if abs(excel_correct - csv_correct) > 0.01 or abs(excel_wrong - csv_wrong) > 0.01:
            errors.append({
                'teacher_id': teacher_id,
                'error': '金额不匹配',
                'excel_correct': excel_correct,
                'excel_wrong': excel_wrong,
                'csv_correct': csv_correct,
                'csv_wrong': csv_wrong,
                'details': details
            })
    
    # 检查是否有Excel中的老师在CSV中缺失
    csv_teachers = set(r['teacher_id'] for r in detail_records)
    for teacher_id in teacher_mapping:
        if teacher_id not in csv_teachers:
            errors.append({
                'teacher_id': teacher_id,
                'error': 'Excel中存在但CSV中缺失',
                'count': len(teacher_mapping[teacher_id])
            })
    
    # 4. 验证汇总文件
    print("\n[4/5] 验证 salary_summary.csv...")
    summary_path = base / "salary_summary.csv"
    csv_summary = {}
    
    with open(summary_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            teacher = row['老师']
            correct = float(row['回答正确所得金'])
            wrong = float(row['回答错误所得金'])
            correct_count = int(row['评价正确题数'])
            wrong_count = int(row['评价错误题数'])
            csv_summary[teacher] = [correct, wrong, correct_count, wrong_count]
    
    print(f"   从 CSV 读取 {len(csv_summary)} 位老师")
    
    # 对比汇总数据
    summary_errors = []
    all_teachers = set(excel_summary.keys()) | set(csv_summary.keys())
    
    for teacher in sorted(all_teachers):
        excel_c, excel_w, excel_cc, excel_wc = excel_summary.get(teacher, [0.0, 0.0, 0, 0])
        csv_c, csv_w, csv_cc, csv_wc = csv_summary.get(teacher, [0.0, 0.0, 0, 0])
        
        # 对比金额和题数
        if (abs(excel_c - csv_c) > 0.01 or abs(excel_w - csv_w) > 0.01 or 
            excel_cc != csv_cc or excel_wc != csv_wc):
            summary_errors.append({
                'teacher': teacher,
                'excel': [excel_c, excel_w, excel_cc, excel_wc],
                'csv': [csv_c, csv_w, csv_cc, csv_wc]
            })
    
    # 5. 输出结果
    print("\n" + "="*80)
    print("验证结果")
    print("="*80)
    
    if not errors and not summary_errors:
        print("\n[OK] 完美！所有记录都匹配，验证通过！")
        print(f"   详细记录验证: {len(detail_records)} 条")
        print(f"   汇总数据验证: {len(csv_summary)} 位老师")
        
        # 显示一些统计信息
        print("\n统计信息:")
        total_correct = sum(r['correct'] for r in detail_records)
        total_wrong = sum(r['wrong'] for r in detail_records)
        total_correct_count = sum(v[2] for v in csv_summary.values())
        total_wrong_count = sum(v[3] for v in csv_summary.values())
        print(f"   总正确金额: {total_correct:.2f}")
        print(f"   总错误金额: {total_wrong:.2f}")
        print(f"   总金额: {total_correct + total_wrong:.2f}")
        print(f"   总正确题数: {total_correct_count}")
        print(f"   总错误题数: {total_wrong_count}")
        print(f"   总题数: {total_correct_count + total_wrong_count}")
    else:
        if errors:
            print(f"\n[ERROR] 详细记录发现 {len(errors)} 个问题：\n")
            for i, err in enumerate(errors, 1):
                print(f"--- 问题 {i} ---")
                print(f"老师ID: {err['teacher_id']}")
                print(f"错误类型: {err['error']}")
                
                if err['error'] == '金额不匹配':
                    print(f"Excel计算: 正确={err['excel_correct']:.2f}, 错误={err['excel_wrong']:.2f}, 合计={err['excel_correct']+err['excel_wrong']:.2f}")
                    print(f"CSV结果:  正确={err['csv_correct']:.2f}, 错误={err['csv_wrong']:.2f}, 合计={err['csv_correct']+err['csv_wrong']:.2f}")
                    print(f"差异: 正确差{err['csv_correct']-err['excel_correct']:.2f}, 错误差{err['csv_wrong']-err['excel_wrong']:.2f}")
                    print("\n详细追踪（Excel原始数据）:")
                    for detail in err['details']:
                        print(f"  {detail}")
                elif err['error'] == 'Excel中存在但CSV中缺失':
                    print(f"该老师在Excel中有 {err['count']} 条记录")
                elif err['error'] == 'CSV中存在但Excel中找不到此老师':
                    print(f"CSV数据: {err['csv_data']}")
                
                print()
        
        if summary_errors:
            print(f"\n[ERROR] 汇总数据发现 {len(summary_errors)} 个问题：\n")
            for i, err in enumerate(summary_errors, 1):
                print(f"--- 问题 {i} ---")
                print(f"老师: {err['teacher']}")
                excel_c, excel_w, excel_cc, excel_wc = err['excel']
                csv_c, csv_w, csv_cc, csv_wc = err['csv']
                print(f"Excel: 正确={excel_c:.2f}({excel_cc}题), 错误={excel_w:.2f}({excel_wc}题), 合计={excel_c+excel_w:.2f}({excel_cc+excel_wc}题)")
                print(f"CSV:   正确={csv_c:.2f}({csv_cc}题), 错误={csv_w:.2f}({csv_wc}题), 合计={csv_c+csv_w:.2f}({csv_cc+csv_wc}题)")
                print()

if __name__ == "__main__":
    main()
